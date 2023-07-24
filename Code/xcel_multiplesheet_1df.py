from openpyxl import load_workbook
import pandas as pd
import datetime
import time
import warnings
warnings.filterwarnings('ignore')
import os

#-----------------------------------------------------------------------------------------------------------------------
## Define the today's date
today = datetime.datetime.today().date()
tday  =today.strftime("%d_%b_%Y")

#-----------------------------------------------------------------------------------------------------------------------

'''Multiple excel sheet of workbookk append into one dataframe'''

## Start
if __name__ == '__main__':
    std_path= r"D:\Test_model/"
    inppath = std_path + "Input/"
    outpth = std_path + "OutPut/" + tday + "/"
    if os.path.exists(outpth):
        pass
    else:
        os.mkdir(outpth)


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

infile = inppath + "Personwise.xlsx"
sheetname = get_sheetnames_xlsx(infile)

lst = []
for i in sheetname:
    if i != "MasterData":
        totalPaidAmount = pd.read_excel(infile, sheet_name=i)
        lst.append(totalPaidAmount)
    else:
        pass
ddfff = pd.DataFrame(pd.concat(lst))

#--------------------------- Final Dump Append
ddfff.to_excel(outpth+"PersonWise_Apended.xlsx",index=False)
