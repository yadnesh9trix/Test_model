import pandas as pd
import numpy as np
from datetime import datetime,timedelta
import warnings
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
import os
import polars


today = datetime.today().date()
last_day = today - timedelta(days=0)

td_date = last_day.strftime("%d")
td_month = last_day.strftime("%m")
last_day_fmt = last_day.strftime("%d%m%Y")

tday  = today.strftime("%d_%b_%Y")

## Start
if __name__ == '__main__':
    main_path = r"D:/"
    std_path= r"D:\Test_model/"
    inppath = std_path + "Input/"
    paidamount_folder = main_path + "Paidamount/"
    outpth = std_path + "OutPut/" + tday + "/"
    if os.path.exists(outpth):
        pass
    else:
        os.mkdir(outpth)

def Ty_paid(paidamount_folder,last_day_fmt):
    typaid = pd.read_excel(paidamount_folder + f"Paidamount_list_{last_day_fmt}.xlsx")
    typaid['propertycode'] = typaid['propertycode'].astype(float)
    typaid['propertycode'] = typaid['propertycode'].apply("{:.02f}".format)
    typaid.dropna(subset=['propertycode'], how='all', inplace=True)

    return typaid

this_year_paid_data = Ty_paid(paidamount_folder,last_day_fmt)

# def get_sheetnames_xlsx(filepath):
#     wb = load_workbook(filepath, read_only=True, keep_links=False)
#     return wb.sheetnames
#
# infile = inppath + "Paid_amountList(Last4Year).xlsx"
# sheetname = get_sheetnames_xlsx(infile)
# lst = []
# for i in sheetname:
#     totalPaidAmount = pd.read_excel(inppath + "Paid_amountList(Last4Year).xlsx", sheet_name=i)
#     lst.append(totalPaidAmount)
# ddfff = pd.DataFrame(pd.concat(lst))

# newwdd = pd.concat([ddfff, this_year_paid_data])
# newwdd = newwdd.sort_values('receiptdate')
#
# newwdd['receiptdate'] = pd.to_datetime(newwdd['receiptdate'], errors='coerce', format='%Y-%m-%d')
# newwdd['fin_year_r'] = newwdd['receiptdate'].dt.year
# newwdd['fin_month_r'] = newwdd['receiptdate'].dt.month
#
# newwdd['fin_year_r'] = np.where(newwdd['fin_month_r'] < 4,
#                                 newwdd['fin_year_r'] - 1,
#                                 newwdd['fin_year_r'])
# newwdd['fin_month_r'] = np.where(newwdd['fin_month_r'] < 4,
#                                  newwdd['fin_month_r'] + 9,
#                                  newwdd['fin_month_r'] - 3)
#
# finYr2018_19 = newwdd[(newwdd['fin_year_r'] >= 2018) & (newwdd['fin_year_r'] < 2019)]
# finYr2019_20 = newwdd[(newwdd['fin_year_r'] >= 2019) & (newwdd['fin_year_r'] < 2020)]
# finYr2020_21 = newwdd[(newwdd['fin_year_r'] >= 2020) & (newwdd['fin_year_r'] < 2021)]
# finYr2021_22 = newwdd[(newwdd['fin_year_r'] >= 2021) & (newwdd['fin_year_r'] < 2022)]
# finYr2022_23 = newwdd[(newwdd['fin_year_r'] >= 2022) & (newwdd['fin_year_r'] < 2023)]
# finYr2023_24 = newwdd[(newwdd['fin_year_r'] >= 2023) & (newwdd['fin_year_r'] < 2024)]

finYr2018_19 = pd.read_csv(inppath + "/paidamount2018-22/Paid_amountlist2018-19.csv")
finYr2018_19['receiptdate'] = pd.to_datetime(finYr2018_19['receiptdate'], errors='coerce', format='%Y-%m-%d')

finYr2019_20 = pd.read_csv(inppath + "/paidamount2018-22/Paid_amountlist2019-20.csv")
finYr2019_20['receiptdate'] = pd.to_datetime(finYr2019_20['receiptdate'], errors='coerce', format='%Y-%m-%d')

finYr2020_21 = pd.read_csv(inppath + "/paidamount2018-22/Paid_amountlist2020-21.csv")
finYr2020_21['receiptdate'] = pd.to_datetime(finYr2020_21['receiptdate'], errors='coerce', format='%Y-%m-%d')

finYr2021_22 = pd.read_csv(inppath + "/paidamount2018-22/Paid_amountlist2021-22.csv")
finYr2021_22['receiptdate'] = pd.to_datetime(finYr2021_22['receiptdate'], errors='coerce', format='%Y-%m-%d')

finYr2022_23 = pd.read_csv(inppath + "/paidamount2018-22/Paid_amountlist2022-23.csv")
finYr2022_23['receiptdate'] = pd.to_datetime(finYr2022_23['receiptdate'], errors='coerce', format='%Y-%m-%d')

finYr2023_24 = this_year_paid_data
finYr2023_24['receiptdate'] = pd.to_datetime(finYr2023_24['receiptdate'], errors='coerce', format='%Y-%m-%d')
finYr2023_24['fin_year_r'] = finYr2023_24['receiptdate'].dt.year
finYr2023_24['fin_month_r'] = finYr2023_24['receiptdate'].dt.month

ddfff = pd.concat([finYr2018_19,finYr2019_20,finYr2020_21,finYr2021_22,finYr2022_23])
newwdd = pd.concat([ddfff, this_year_paid_data])
newwdd = newwdd.sort_values('receiptdate')

newwdd['receiptdate'] = pd.to_datetime(newwdd['receiptdate'], errors='coerce', format='%Y-%m-%d')
newwdd['fin_year_r'] = newwdd['receiptdate'].dt.year
newwdd['fin_month_r'] = newwdd['receiptdate'].dt.month

newwdd['fin_year_r'] = np.where(newwdd['fin_month_r'] < 4,
                                newwdd['fin_year_r'] - 1,
                                newwdd['fin_year_r'])
newwdd['fin_month_r'] = np.where(newwdd['fin_month_r'] < 4,
                                 newwdd['fin_month_r'] + 9,
                                 newwdd['fin_month_r'] - 3)

finYr2018_19 = newwdd[(newwdd['fin_year_r'] >= 2018) & (newwdd['fin_year_r'] < 2019)]
finYr2019_20 = newwdd[(newwdd['fin_year_r'] >= 2019) & (newwdd['fin_year_r'] < 2020)]
finYr2020_21 = newwdd[(newwdd['fin_year_r'] >= 2020) & (newwdd['fin_year_r'] < 2021)]
finYr2021_22 = newwdd[(newwdd['fin_year_r'] >= 2021) & (newwdd['fin_year_r'] < 2022)]
finYr2022_23 = newwdd[(newwdd['fin_year_r'] >= 2022) & (newwdd['fin_year_r'] < 2023)]
finYr2023_24 = newwdd[(newwdd['fin_year_r'] >= 2023) & (newwdd['fin_year_r'] < 2024)]


df = pd.concat([finYr2022_23,finYr2023_24],ignore_index=False)
grpbydf = df.groupby(['receiptdate']).agg(
    {'propertycode': 'count', 'paidamount': 'sum'}).reset_index()
grpbydf_Last_year = grpbydf[(grpbydf['receiptdate'] >= datetime(2022, 6, 1)) & (grpbydf['receiptdate'] <= datetime(2022, 6, int(td_date)))]
grpbydf_Last_year = grpbydf_Last_year.rename(columns= {'paidamount':'paidamount_ly','propertycode':'propertycode_ly'}).reset_index(drop=True)

grpbydf_This_year = grpbydf[grpbydf['receiptdate'] >= datetime(2023, 6, 1)].reset_index(drop=True)
grpbydf_This_year = grpbydf_This_year.rename(columns= {'paidamount':'paidamount_ty','propertycode':'propertycode_ty'}).reset_index(drop=True)
grpbydf_This_year[['propertycode_ly','paidamount_ly']] = grpbydf_Last_year[['propertycode_ly','paidamount_ly']]

grpbydf_This_year[['paidamount_ty','paidamount_ly']] = grpbydf_This_year[['paidamount_ty','paidamount_ly']]/10000000

grpbydf_This_year.to_excel(outpth + "Ly_Ty_Collection.xlsx",index=False)

##-----------------------------------------------------------------------------------------------------------------------
# grpbydf['month'] = grpbydf['receiptdate'].dt.strftime("%b")
# grpbydf_Last_year1 = grpbydf[(grpbydf['receiptdate'] >= datetime(2022, 4, 1)) & (grpbydf['receiptdate'] <= datetime(2022, 6, int(td_date)))]
# grpbydf_Last_year1 = grpbydf_Last_year1.drop(columns='receiptdate')
# grpbydf_Last_year1 = grpbydf_Last_year1.rename(columns= {'paidamount':'paidamount_ly','propertycode':'propertycode_ly'}).reset_index(drop=True)
# month_grpbydf_ly = grpbydf_Last_year1.groupby(['month']).agg(
#     {'paidamount_ly': 'sum', 'propertycode_ly': 'sum'}).reset_index()
#
# grpbydf_this_year1 = grpbydf[(grpbydf['receiptdate'] >= datetime(2023,4,1))]
# grpbydf_this_year1 = grpbydf_this_year1.drop(columns='receiptdate')
# grpbydf_this_year1 = grpbydf_this_year1.rename(columns= {'paidamount':'paidamount_ty','propertycode':'propertycode_ty'}).reset_index(drop=True)
# month_grpbydf_ty = grpbydf_this_year1.groupby(['month']).agg(
#     {'paidamount_ty': 'sum', 'propertycode_ty': 'sum'}).reset_index()
# month_grpbydf_ty[['propertycode_ly','paidamount_ly']] = month_grpbydf_ly[['propertycode_ly','paidamount_ly']]
#
# month_grpbydf_ty.to_excel(outpth + "Ly_Ty_monthwise_Collection.xlsx",index=False)
#-----------------------------------------------------------------------------------------------------------------------

yr = [2018, 2019, 2020, 2021, 2022, 2023]
lll = []
for i in yr:
    finYr2023_24 = newwdd[(newwdd['fin_year_r'] >= i) & (newwdd['fin_year_r'] < i + 1)]
    df_filtered = finYr2023_24[finYr2023_24['receiptdate'] <= datetime(i, int(td_month), int(td_date))]
    df_filtered['Financial_Year'] = f"{i}-{i + 1}"
    df_filtered['STLY_Date'] = datetime(i, int(td_month), int(td_date))
    grpbydf = df_filtered.groupby(['Financial_Year', 'ezname', 'STLY_Date']).agg(
        {'propertycode': 'count', 'paidamount': 'sum'}).reset_index()
    lll.append(grpbydf)

ddd1 = pd.DataFrame(pd.concat(lll))

ddd1.to_excel(outpth + "L5Y_Collection.xlsx", index=False)
